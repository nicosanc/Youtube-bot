from googleapiclient.discovery import build
import datetime
import pandas as pd
from googleapiclient.http import MediaFileUpload
import os
from app.services.auth_service import load_credentials
from app.config import settings
from openpyxl import load_workbook

def get_or_create_daily_folder(drive, parent_folder_id: str) -> str:
    """Find or create today's folder (format: mm-dd-yyyy)"""
    today = datetime.datetime.now().strftime('%m-%d-%Y')
    
    # Search for today's folder
    query = f"name='{today}' and '{parent_folder_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = drive.files().list(q=query, fields='files(id, name)').execute()
    folders = results.get('files', [])
    
    if folders:
        return folders[0]['id']
    
    # Create folder if it doesn't exist
    folder_metadata = {
        'name': today,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_folder_id]
    }
    folder = drive.files().create(body=folder_metadata, fields='id').execute()
    return folder['id']

def get_next_task_number(drive, folder_id: str) -> int:
    """Count existing files in folder and return next task number"""
    query = f"'{folder_id}' in parents and trashed=false"
    results = drive.files().list(q=query, fields='files(id)').execute()
    files = results.get('files', [])
    return len(files) + 1

def send_excel_to_drive(data: dict, user_email: str) -> str:
    # Load user's OAuth credentials
    credentials = load_credentials(user_email)
    if not credentials:
        raise Exception("User not authenticated. Please login first.")
    
    # Build Drive service with user's credentials
    drive = build('drive', 'v3', credentials=credentials)
    
    # Get or create today's folder
    daily_folder_id = get_or_create_daily_folder(drive, settings.GOOGLE_DRIVE_FOLDER_ID)
    
    # Get next task number
    task_number = get_next_task_number(drive, daily_folder_id)
    
    # Create DataFrame and Excel file
    df = pd.DataFrame(data)
    filename = f"Task {task_number} Output.xlsx"
    filepath = f"/tmp/{filename}"
    df.to_excel(filepath, index=False)
    
    # Add CPM formula column using openpyxl
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Add CPM header in the next column
    last_col = ws.max_column + 1
    ws.cell(row=1, column=last_col, value='CPM')
    
    # Add CPM formula for each data row (formula: Price / (Avg Views / 1000))
    # Price is in column J (10th column), Avg Views is in column B (2nd column)
    for row in range(2, ws.max_row + 1):
        price_col = chr(64 + 10)  # J
        views_col = chr(64 + 2)   # B (Avg Views)
        formula = f'=IF({price_col}{row}=0, "", {price_col}{row}/({views_col}{row}/1000))'
        ws.cell(row=row, column=last_col, value=formula)
    
    wb.save(filepath)
    
    # Upload to today's folder (not main folder)
    file_metadata = {
        'name': filename,
        'parents': [daily_folder_id]
    }
    
    media = MediaFileUpload(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file = drive.files().create(body=file_metadata, media_body=media, fields='id').execute()
    os.remove(filepath)
    
    print(f"File uploaded: {filename} to folder {daily_folder_id}")
    
    return f"https://drive.google.com/file/d/{file.get('id')}/view"